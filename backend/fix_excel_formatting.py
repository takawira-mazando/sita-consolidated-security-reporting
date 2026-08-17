from copy import copy

import openpyxl

file_path = r"C:\Users\HosiTech\Downloads\HOSI-UNIFIED DATA MANAGED SERVICE PROJECT FLIGHT PLAN 2026.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["Project Plan "]

# We want to remove the blank row 79
ws.delete_rows(79, 1)

# Now what was at 80 is at 79, etc.
# 79: Frontend SAST
# 80: Vulnerability Remediation
# 81: Tenancy Hierarchical
# 82: Delegated RBAC

# Insert 1 row at 79 for Backend SAST
ws.insert_rows(79, 1)

# Set the values for Backend SAST
ws.cell(row=79, column=1, value="")
ws.cell(row=79, column=2, value="")
ws.cell(row=79, column=3, value="Backend SAST Setup & Scan (Bandit)")
ws.cell(row=79, column=4, value="Completed")
ws.cell(row=79, column=5, value="1")

# The other rows are now at 80, 81, 82, 83
# Let's copy styling from row 74 (a standard task row) to rows 79..83
source_row = 74

for r in range(79, 84):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=r, column=col)
        
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

wb.save(file_path)
print("Excel file successfully updated: Added Backend SAST and standardised formatting.")
