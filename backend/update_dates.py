import datetime

import openpyxl

file_path = r"C:\Users\HosiTech\Downloads\HOSI-UNIFIED DATA MANAGED SERVICE PROJECT FLIGHT PLAN 2026.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["Project Plan "]

start_date = datetime.datetime(2026, 8, 11)
end_date = datetime.datetime(2026, 8, 12)

for r in range(79, 91):
    status_cell = ws.cell(row=r, column=4)
    if status_cell.value == "Completed":
        # Percentage must be 100% (value = 1, format = '0%')
        ws.cell(row=r, column=5, value=1)
        ws.cell(row=r, column=5).number_format = '0%'
        
        # Add Dates
        ws.cell(row=r, column=6, value=start_date)
        ws.cell(row=r, column=6).number_format = 'd-mmm-yy'
        
        ws.cell(row=r, column=7, value=end_date)
        ws.cell(row=r, column=7).number_format = 'd-mmm-yy'
        
        # Add Assignee
        ws.cell(row=r, column=8, value='Lead Dev')

wb.save(file_path)
print("Excel file fully updated with 100% completion and dates.")
