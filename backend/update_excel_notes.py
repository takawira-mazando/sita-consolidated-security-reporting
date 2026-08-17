import openpyxl

file_path = r"C:\Users\HosiTech\Downloads\HOSI-UNIFIED DATA MANAGED SERVICE PROJECT FLIGHT PLAN 2026.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["Project Plan "]

for r in range(79, 90):
    task_name = ws.cell(row=r, column=3).value
    if task_name and "Backend SAST Setup & Scan" in task_name:
        ws.cell(row=r, column=9, value="Scan complete: 8222 lines checked, 0 issues found. Report generated.")
        break

wb.save(file_path)
print("Excel file successfully updated with Bandit scan report notes.")
