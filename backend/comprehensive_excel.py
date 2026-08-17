from copy import copy

import openpyxl
from openpyxl.styles import PatternFill

file_path = r"C:\Users\HosiTech\Downloads\HOSI-UNIFIED DATA MANAGED SERVICE PROJECT FLIGHT PLAN 2026.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["Project Plan "]

# Delete rows 79 and downwards to clean up any previous mess
max_r = ws.max_row
if max_r >= 79:
    ws.delete_rows(79, max_r - 79 + 1)

tasks_to_add = [
    # SAST
    ['', '', 'Backend SAST Setup & Scan (Bandit)', 'Completed', '1'],
    ['', '', 'Frontend SAST Setup & Scan (Semgrep)', 'Completed', '1'],
    ['', '', 'Vulnerability Remediation', 'Completed', '1'],
    # Comprehensive Tenancy
    ['', '', 'Tenancy Data Model (Provinces, Departments, Branches)', 'Completed', '1'],
    ['', '', 'API-Level Tenant Filtering & Data Isolation (RLS)', 'Completed', '1'],
    ['', '', 'Tenancy-Aware JWT Authentication Scopes', 'Completed', '1'],
    ['', '', 'Delegated Role-Based Access Control (RBAC) Logic', 'Completed', '1'],
    ['', '', 'Frontend Tenant Context & Dynamic UI Layouts', 'Completed', '1'],
    ['', '', 'Frontend Scoped User Management UI (Admin Users)', 'Completed', '1'],
    ['', '', 'Cross-Tenant Access Prevention & Security Auditing', 'Completed', '1'],
    ['', '', 'Multi-Tenancy E2E Testing & Verification', 'Completed', '1'],
    ['', '', 'Tenancy Architecture Documentation', 'Completed', '1'],
    # Phase 12 header
    ['', 'Phase 12: UAT, Training & Go-Live', 'User acceptance testing', 'Not started', ''],
    # Phase 12 tasks
    ['', '', 'Runbook & operations guide', 'Not started', ''],
    ['', '', 'Stakeholder demo & training', 'Not started', ''],
    ['', '', 'Security review & penetration test', 'Not started', ''],
    ['', '', 'Production go-live', 'Not started', ''],
    ['', '', 'Risk: Compliance data migration incomplete', 'WIP', ''],
]

for idx, data in enumerate(tasks_to_add):
    r = 79 + idx
    for c_idx, val in enumerate(data):
        ws.cell(row=r, column=c_idx+1, value=val)
        
    is_header = "Phase 12:" in str(data[1])
    is_wip = "WIP" in str(data[3])
    is_not_started = "Not started" in str(data[3])
    
    if is_header:
        ref_row = 73
    elif is_wip:
        ref_row = 78
    else:
        ref_row = 74 # Default to Completed task style
        
    for col in range(1, 10): 
        src_cell = ws.cell(row=ref_row, column=col)
        tgt_cell = ws.cell(row=r, column=col)
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.protection = copy(src_cell.protection)
            tgt_cell.alignment = copy(src_cell.alignment)

red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
for idx, data in enumerate(tasks_to_add):
    r = 79 + idx
    if "Not started" in str(data[3]):
        ws.cell(row=r, column=4).fill = red_fill
    if "WIP" in str(data[3]):
        ws.cell(row=r, column=4).fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    if "Completed" in str(data[3]) and "Phase 12:" not in str(data[1]):
        ws.cell(row=r, column=4).fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")

wb.save(file_path)
print("Excel file fully updated with comprehensive tenancy tasks and standardized formatting.")
