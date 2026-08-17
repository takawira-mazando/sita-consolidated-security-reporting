import openpyxl

file_path = r"C:\Users\HosiTech\Downloads\HOSI-UNIFIED DATA MANAGED SERVICE PROJECT FLIGHT PLAN 2026.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["Project Plan "]

# Rows where the deliverables currently reside
rows_to_extract = [89, 90, 91, 94, 95]
extracted_data = []
for r in rows_to_extract:
    extracted_data.append([cell.value for cell in ws[r]])

# Clean up the appended phases (from row 87 downwards)
max_r = ws.max_row
if max_r >= 87:
    ws.delete_rows(87, max_r - 87 + 1)

# Insert 5 rows at row 79 (end of Phase 11)
ws.insert_rows(79, 5)

# Write the deliverables into the new rows
for i, data in enumerate(extracted_data):
    for j, val in enumerate(data):
        ws.cell(row=79+i, column=j+1, value=val)

wb.save(file_path)
print("Excel file successfully updated: deliverables moved to Phase 11.")
