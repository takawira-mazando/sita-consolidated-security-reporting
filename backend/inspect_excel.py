import openpyxl

file_path = r"C:\Users\HosiTech\Downloads\HOSI-UNIFIED DATA MANAGED SERVICE PROJECT FLIGHT PLAN 2026.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["Project Plan "]

max_row = ws.max_row
with open("excel_dump.txt", "w", encoding="utf-8") as f:
    f.write(f"Max row: {max_row}\n")
    for i in range(max_row - 20, max_row + 1):
        row_data = []
        for cell in ws[i]:
            row_data.append(str(cell.value).replace('\n', ' ')[:30] if cell.value else "")
        f.write(f"Row {i}: {row_data}\n")
